import serial
import pyautogui
# import time
# import subprocess
from openpyxl import load_workbook

ser = serial.Serial('COM3', baudrate=115200)

FILE_PATH = 'serial_data.txt'
FILE_PATH1 = 'MAC_List_test.xlsx'
PRINTER_PATH = r'C:\Program Files\Zebra Technologies\ZebraDesigner 3\bin.net\ZebraDesigner.exe'

# Location of QR and SN fields in printing window of ZebraDesigner3
# X_QR_COORDINATE = 230
# Y_QR_COORDINATE = 487
#
# X_SN_COORDINATE = 224
# Y_SN_COORDINATE = 516
#
# X_BUTTON_PRINT_COORDINATE = 237
# Y_BUTTON_PRINT_COORDINATE = 98

################## Lenovo laptop coordinates ##################
X_QR_COORDINATE = 519
Y_QR_COORDINATE = 616

X_SN_COORDINATE = 519
Y_SN_COORDINATE = 666

X_BUTTON_PRINT_COORDINATE = 300
Y_BUTTON_PRINT_COORDINATE = 120

# Variable to store extracted MAC address and DUID from .txt file, where all tag data is stored
extractedData = []


# NFC tag reading and writing to txt file
def read_card(file_path):
    line_data = ''
    with open(file_path, 'w') as file:
        while line_data != 'Operation completed':
            # for data in file_path:
            # if ser.in_waiting > 0:
            line_data = ser.readline().decode('utf-8').strip()
            print(line_data)
            file.write(line_data + '\n')
        file.flush()
        file.close()
        print("File closed.")


# MAC and DUID extracted from txt file and data is stored to list collectedData
def read_block(file_path):
    # mac_address = []
    # duid_data = []

    with open(file_path, 'r') as file:
        for line in file:
            if line.startswith('MAC'):
                mac_address = line.split(':')
                value = mac_address[1].strip()
                extractedData.insert(0, value)
            elif line.startswith('DUID'):
                duid_data = line.split(':')
                value1 = duid_data[1].strip()
                extractedData.insert(1, value1)
    print(extractedData)


# Access mac address list on Excel
def check_values(file_path):
    company_name = 'VIEZO'
    value_found = False
    workbook = load_workbook(file_path)
    # List all sheet names in workbook
    # sheet_names = workbook.sheetnames
    second_sheet = workbook.worksheets[1]
    print(second_sheet)

    # Column to check MAC address for dictionary values
    column_to_check = 'D'
    # Iterate over cells in the column
    for cell in second_sheet[column_to_check]:
        if cell.value == extractedData[0]:
            duid_value = cell.offset(column=4).value
            if duid_value == extractedData[1]:
                serial_number_cell = cell.offset(column=6)
                serial_number = generate_serial(company_name, duid_value)
                serial_number_cell.value = serial_number
                extractedData.insert(2, serial_number_cell.value)
                workbook.save(FILE_PATH1)
                print(serial_number_cell.value)
                print('Sensor exist in MAC address list')
                print_by_zebra(extractedData)
                extractedData.clear()
                value_found = True
                break

    # if value is not found, add values from extractedData to the first empty cell of column
    if not value_found:
        for cell in second_sheet[column_to_check]:
            if cell.value is None or cell.value == "":
                try:
                    cell.value = extractedData[0]
                    print(cell.value)
                    cell.offset(column=4).value = extractedData[1]
                    print(cell.offset(column=4).value)

                    # checking to avoid that MAC address is the same as DUID and vice verse
                    if extractedData[0] != extractedData[1]:
                        serial_number_cell = cell.offset(column=6)
                        serial_number = generate_serial(company_name, cell.offset(column=4).value)
                        serial_number_cell.value = serial_number
                        extractedData.append(serial_number_cell.value)
                        workbook.save(FILE_PATH1)
                        print('Values added to empty spaces in the particular columns')
                        print_by_zebra(extractedData)
                        extractedData.clear()
                except Exception as e:
                    print(f"Error occurred while adding data: {e}")
                    extractedData.clear()
                break


def print_by_zebra(data):
    # Mouse pointer to select and delete any old information from QR field
    pyautogui.click(x=X_QR_COORDINATE, y=Y_QR_COORDINATE)

    # Mouse pointer to select and delete any old information from SN (serial number) field
    pyautogui.hotkey('ctrl', 'a')  # Select all
    pyautogui.press('delete')  # Delete selected text

    pyautogui.click(x=X_SN_COORDINATE, y=Y_SN_COORDINATE)

    # Send a keyboard shortcut for the delete action
    pyautogui.hotkey('ctrl', 'a')  # Select all
    pyautogui.press('delete')  # Delete selected text

    # Pass new sensor data to be printed
    pyautogui.click(x=X_QR_COORDINATE, y=Y_QR_COORDINATE)
    # pyautogui.typewrite(str(data[0]))
    serial_number = f"S/N: {data[2]}"
    pyautogui.typewrite(serial_number)
    pyautogui.press('enter')

    # pyautogui.click(x=X_SN_COORDINATE, y=Y_SN_COORDINATE)
    pyautogui.typewrite(str(data[2]))
    pyautogui.press('enter')

    # Press Print button on Zebra Designer software
    pyautogui.click(x=X_BUTTON_PRINT_COORDINATE, y=Y_BUTTON_PRINT_COORDINATE)


#  -----------------------------------------------------------------------------
def pointerPosition():
    # Get the current mouse pointer coordinates
    x, y = pyautogui.position()

    # Print the coordinates
   # print(f"Mouse pointer coordinates: ({x}, {y})")
# ------------------------------------------------------------------------------

def generate_serial(company_name, duid_value):
    serial_no = str(duid_value[:])
    return serial_no


def main():
    pointerPosition()
    while 1:
        read_card(FILE_PATH)
        read_block(FILE_PATH)
        check_values(FILE_PATH1)


if __name__ == "__main__":
    main()
